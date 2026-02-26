#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
supp_table11.py
==============
Supplementary Table S11 (SZ only):
Clinical associations with clustered ripple state metrics (pooled 80–240 Hz)

Purpose
-------
This script produces Supplementary Table S11, summarizing associations between
clinical measures and subject-level metrics of temporally clustered (high-rate)
ripple epochs within schizophrenia participants.

Modeling framework (aligned with S10/S11 policy)
----------------------------------------------
- One model per clinical predictor (PANSS POS / NEG / GEN / GAF) per outcome.
- All models additionally adjust for:
    * total ripple load (events_sum)
    * demographic/clinical covariates (age, sex, JART, sleepiness_pre, antipsychotics)
    * recording site (C(site_public))
- Count outcomes:
    * Negative Binomial GLM with log link
    * offset = log(minutes_sum)
    * HC3 robust standard errors
    * effect size: IRR = exp(beta) with 95% CI
- Continuous outcome:
    * OLS with HC3 robust standard errors
    * effect size: beta with 95% CI
- Multiple comparisons:
    * BH-FDR across clinical predictors WITHIN each outcome

Outcomes (pooled across 80–240 Hz)
----------------------------------
(i)   clustered_epochs
      = sum(n_epochs) across frequencies
(ii)  outside_ripple_count
      = sum(n_events_outside_epochs) across frequencies
(iii) within_epoch_rate (continuous)
      = duration-weighted mean across frequencies:
        weighted_mean(within_epoch_event_rate, weights=sum_epoch_dur_s)

Definition note
---------------
within_epoch_event_rate is defined upstream as:
    n_events_in_epochs / sum_epoch_dur_s
(duration-weighted within-state density; NOT mean-of-epoch-densities)

Inputs (under <root>/data/)
--------------------------
- public_clinical_subject_level.csv      (SAFE; anonymized)
- anon_id_map.csv (preferred) OR anon_id_map.csv (legacy; subject <-> anon_id mapping)
- epoch_metrics_subject_freq.csv        (derived table from temporal clustering pipeline)

Outputs (under <root>/results/tables/)
--------------------------------------
- Table_S11_clinical_associations_with_cluster_metrics.csv
- (optional) Supplementary_Table_S11_clinical_associations_with_cluster_metrics.xlsx
  written only if a template is available (supp_table11.xlsx)

Reproducibility / Code Ocean
----------------------------
- --root CLI argument OR env var PROJECT_ROOT
- fallback: auto-detect repo root from CWD by searching for run_all.py
- no absolute paths
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
import re
import numpy as np
import pandas as pd

import openpyxl
import statsmodels.api as sm
import statsmodels.formula.api as smf
from statsmodels.stats.multitest import multipletests


# -------------------------
# Configuration
# -------------------------
FREQS_USE = [80, 120, 160, 200, 240]
COVARS = ["age", "sex", "JART", "sleepiness_pre", "antipsychotics"]

# Predictors (one model per predictor)
# NOTE: legacy project name "PANSS_pasological" is supported as alias for PANSS_general.
PRED_MAP = [
    ("PANSS_positive", "PANSS positive"),
    ("PANSS_negative", "PANSS negative"),
    ("PANSS_pasological", "PANSS general"),
    ("GAF", "GAF"),
]

OUTCOME_CONT = "within_epoch_rate"  # keep consistent with manuscript wording


# -------------------------
# Root/path helpers
# -------------------------
def _guess_repo_root(start: Path) -> Path:
    start = start.resolve()
    for p in [start] + list(start.parents):
        if (p / "run_all.py").exists():
            return p
    for p in [start] + list(start.parents):
        if (p / "data").exists():
            return p
    return start

def get_root(cli_root: str | None = None) -> Path:
    if cli_root:
        return Path(cli_root).expanduser().resolve()
    env = os.environ.get("PROJECT_ROOT", "").strip()
    if env:
        return Path(env).expanduser().resolve()
    return _guess_repo_root(Path.cwd())

def _find_template(root: Path, name: str) -> Path | None:
    candidates = [
        root / "templates" / name,
        root / "template" / name,
        root / "supp_tables" / name,
        Path(__file__).resolve().parent / name,
        Path(__file__).resolve().parent.parent / name,
    ]
    for c in candidates:
        if c.exists():
            return c
    return None


# -------------------------
# Utilities
# -------------------------
def canon_subject(x) -> str:
    s = str(x).strip()
    m = re.search(r"(\d+)", s)
    return f"NB_subject_{int(m.group(1))}" if m else s

def _resolve_panss_gen(df: pd.DataFrame) -> pd.DataFrame:
    """
    Backward compatibility:
    - If PANSS_pasological is missing but PANSS_general exists, copy it.
    """
    df = df.copy()
    if "PANSS_pasological" not in df.columns and "PANSS_general" in df.columns:
        df["PANSS_pasological"] = df["PANSS_general"]
    return df

def _dropna(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    return df.replace([np.inf, -np.inf], np.nan).dropna(subset=cols, how="any").copy()

def _ensure_numeric(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


# -------------------------
# Optional XLSX writer (template-based)
# -------------------------
def write_xlsx_from_template(df_out: pd.DataFrame, template_path: Path, out_path: Path) -> None:
    """
    Template expected layout (typical):
      columns A..H:
        Outcome | Predictor | Effect_type | Effect | CI_low | CI_high | p | q(FDR)
    Data rows start at row 3.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    # clear data area generously
    for r in range(3, 300):
        for c in range(1, 9):
            ws.cell(r, c).value = None

    r0 = 3
    for i, row in df_out.iterrows():
        ws.cell(r0 + i, 1).value = str(row["outcome_label"])
        ws.cell(r0 + i, 2).value = str(row["predictor_label"])
        ws.cell(r0 + i, 3).value = str(row["effect_type"])
        ws.cell(r0 + i, 4).value = float(row["effect"]) if np.isfinite(row["effect"]) else None
        ws.cell(r0 + i, 5).value = float(row["ci_low"]) if np.isfinite(row["ci_low"]) else None
        ws.cell(r0 + i, 6).value = float(row["ci_high"]) if np.isfinite(row["ci_high"]) else None
        ws.cell(r0 + i, 7).value = float(row["p_value"]) if np.isfinite(row["p_value"]) else None
        ws.cell(r0 + i, 8).value = float(row["q_fdr"]) if np.isfinite(row["q_fdr"]) else None

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def find_root() -> Path:
    start = Path.cwd().resolve()
    try:
        start = Path(__file__).resolve()
    except NameError:
        pass
    for p in [start] + list(start.parents):
        if (p / "run_all.py").exists():
            return p
    for p in [start] + list(start.parents):
        if (p / "data").exists():
            return p
    return Path.cwd().resolve()
    
# -------------------------
# Main
# -------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", type=str, default=None, help="Project root (default: env PROJECT_ROOT or auto-detect)")
    ap.add_argument("--template", type=str, default=None, help="Optional template xlsx (supp_table11.xlsx)")
    ap.add_argument("--id-map", type=str, default=None,
                    help="Optional id map path. Default: data/anon_id_map.csv (fallback: data/anon_id_map.csv)")
    args = ap.parse_args()

    root = Path(args.root).expanduser().resolve() if args.root else find_root()
    data_dir = root / "data"
    if Path("/results").exists():
        results_root = Path("/results")
    else:
        results_root = root / "results"
    out_dir = results_root / "tables"
    out_dir.mkdir(parents=True, exist_ok=True)


    # Inputs
    clin_csv = data_dir / "public_clinical_subject_level.csv"
    rate_csv = data_dir / "epoch_metrics_subject_freq.csv"

    # ID map: prefer anon_id_map.csv, fallback to legacy anon_id_map.csv
    if args.id_map:
        id_map = Path(args.id_map).expanduser().resolve()
    else:
        id_map = data_dir / "anon_id_map.csv"
        if not id_map.exists():
            id_map = data_dir / "private_anon_id_map.csv"

    for p in [clin_csv, id_map, rate_csv]:
        if not p.exists():
            raise FileNotFoundError(f"Missing required input: {p}")

    # ---- clinical (SAFE) ----
    clin = pd.read_csv(clin_csv)
    clin = _resolve_panss_gen(clin)

    req = {"anon_id", "site_public", "events_sum", "minutes_sum"}
    miss = req - set(clin.columns)
    if miss:
        raise ValueError(f"{clin_csv.name} missing required columns: {sorted(miss)}")

    num_cols = ["events_sum", "minutes_sum"] + COVARS + [c for c, _ in PRED_MAP]
    clin = _ensure_numeric(clin, [c for c in num_cols if c in clin.columns])
    clin["site_public"] = clin["site_public"].astype(str)
    clin["anon_id"] = clin["anon_id"].astype(str)

    # SZ-only filter if group exists (recommended)
    if "group" in clin.columns:
        clin["group"] = clin["group"].astype(str).str.upper().replace({"SC": "SZ", "SCZ": "SZ", "SCHIZOPHRENIA": "SZ"})
        clin = clin[clin["group"].isin(["SZ"])].copy()

    # ---- id map ----
    mp = pd.read_csv(id_map)
    if not {"subject", "anon_id"}.issubset(mp.columns):
        raise ValueError(f"{id_map.name} must have columns: subject, anon_id")
    mp = mp.copy()
    mp["subject"] = mp["subject"].map(canon_subject)
    mp["anon_id"] = mp["anon_id"].astype(str)

    # ---- epoch metrics (subject × freq) ----
    rate = pd.read_csv(rate_csv)
    need_rate = {
        "subject", "freq",
        "n_epochs", "n_events", "n_events_in_epochs",
        "sum_epoch_dur_s",
        "within_epoch_event_rate",
    }
    miss = need_rate - set(rate.columns)
    if miss:
        raise ValueError(
            f"{rate_csv.name} missing required columns: {sorted(miss)}\n"
            "Expected: within_epoch_event_rate = n_events_in_epochs / sum_epoch_dur_s"
        )

    rate = rate.copy()
    rate["subject"] = rate["subject"].map(canon_subject)
    rate["freq"] = pd.to_numeric(rate["freq"], errors="coerce").astype(int)
    rate = _ensure_numeric(rate, [
        "n_epochs", "n_events", "n_events_in_epochs", "sum_epoch_dur_s", "within_epoch_event_rate"
    ])

    rate = rate[rate["freq"].isin(FREQS_USE)].copy()
    rate["n_events_outside_epochs"] = rate["n_events"] - rate["n_events_in_epochs"]

    # ---- pool outcomes across frequencies (NO groupby.apply; pandas-stable) ----
    w = pd.to_numeric(rate["sum_epoch_dur_s"], errors="coerce").to_numpy(float)
    y = pd.to_numeric(rate["within_epoch_event_rate"], errors="coerce").to_numpy(float)
    good = np.isfinite(w) & np.isfinite(y) & (w > 0)

    rate["_w"] = np.where(good, w, 0.0)
    rate["_wy"] = np.where(good, w * y, 0.0)

    agg = (rate.groupby("subject", as_index=False)
              .agg(
                  clustered_epochs=("n_epochs", "sum"),
                  outside_ripple_count=("n_events_outside_epochs", "sum"),
                  _w_sum=("_w", "sum"),
                  _wy_sum=("_wy", "sum"),
              ))

    agg["within_epoch_rate"] = np.where(
        agg["_w_sum"] > 0,
        agg["_wy_sum"] / agg["_w_sum"],
        np.nan
    )
    agg = agg.drop(columns=["_w_sum", "_wy_sum"])

    # Attach anon_id via mapping
    agg = agg.merge(mp, on="subject", how="left").dropna(subset=["anon_id"]).copy()
    agg["anon_id"] = agg["anon_id"].astype(str)

    # Merge to clinical
    D = clin.merge(agg, on="anon_id", how="inner").replace([np.inf, -np.inf], np.nan)

    outcomes = [
        ("clustered_epochs", "Number of high-rate ripple epochs (pooled 80–240 Hz)", "count", "IRR"),
        ("outside_ripple_count", "Ripple events outside high-rate epochs (pooled 80–240 Hz)", "count", "IRR"),
        ("within_epoch_rate", "Within-epoch ripple rate (events/s; pooled 80–240 Hz)", "cont", "beta"),
    ]

    # NB family for counts (alpha explicitly set to avoid warnings)
    fam = sm.families.NegativeBinomial(alpha=1.0, link=sm.families.links.Log())

    rows = []
    for out_col, out_label, out_type, effect_type in outcomes:
        tmp = []
        for pred_col, pred_label in PRED_MAP:
            if pred_col not in D.columns:
                continue

            covars_use = [c for c in COVARS if c in D.columns]

            if out_type == "count":
                cols_need = ["site_public", "minutes_sum", "events_sum", out_col, pred_col] + covars_use
            else:
                cols_need = ["site_public", "events_sum", out_col, pred_col] + covars_use

            X = _dropna(D, cols_need)
            if X.empty:
                continue

            if out_type == "count":
                X = X[pd.to_numeric(X["minutes_sum"], errors="coerce") > 0].copy()
                if X.empty:
                    continue

                formula = f"{out_col} ~ {pred_col} + events_sum"
                if covars_use:
                    formula += " + " + " + ".join(covars_use)
                formula += " + C(site_public)"

                fit = smf.glm(
                    formula=formula,
                    data=X,
                    family=fam,
                    offset=np.log(X["minutes_sum"].astype(float))
                ).fit(cov_type="HC3")

                beta = float(fit.params[pred_col])
                se = float(fit.bse[pred_col])
                p = float(fit.pvalues[pred_col])

                eff = float(np.exp(beta))
                ci_low = float(np.exp(beta - 1.96 * se))
                ci_high = float(np.exp(beta + 1.96 * se))

            else:
                formula = f"{out_col} ~ {pred_col} + events_sum"
                if covars_use:
                    formula += " + " + " + ".join(covars_use)
                formula += " + C(site_public)"

                fit = smf.ols(formula=formula, data=X).fit(cov_type="HC3")

                beta = float(fit.params[pred_col])
                se = float(fit.bse[pred_col])
                p = float(fit.pvalues[pred_col])

                eff = beta
                ci_low = beta - 1.96 * se
                ci_high = beta + 1.96 * se

            tmp.append(dict(
                outcome=out_col,
                outcome_label=out_label,
                predictor=pred_col,
                predictor_label=pred_label,
                effect_type=effect_type,
                effect=eff,
                ci_low=ci_low,
                ci_high=ci_high,
                p_value=p,
                n=int(X.shape[0]),
            ))

        if tmp:
            df_tmp = pd.DataFrame(tmp)
            df_tmp["q_fdr"] = multipletests(df_tmp["p_value"].values, method="fdr_bh")[1]
            rows.append(df_tmp)

    if not rows:
        raise RuntimeError("No models were fit. Check predictors/inputs and NA filtering.")

    out = pd.concat(rows, ignore_index=True)

    out_csv = out_dir / "Table_S11_clinical_associations_with_cluster_metrics.csv"
    out.to_csv(out_csv, index=False)

    # Optional XLSX
    tpl = None
    if args.template:
        tp = Path(args.template).expanduser().resolve()
        tpl = tp if tp.exists() else None
    if tpl is None:
        tpl = _find_template(root, "supp_table11.xlsx")

    if tpl is not None:
        out_xlsx = out_dir / "Supplementary_Table_S11_clinical_associations_with_cluster_metrics.xlsx"
        try:
            write_xlsx_from_template(out, tpl, out_xlsx)
        except Exception as e:
            print("[WARN] Failed to write XLSX from template:", repr(e))

    print("[OK] root:", root)
    print("[OK] id_map:", id_map)
    print("[OK] wrote:", out_csv)
    print("[INFO] pooled freqs:", FREQS_USE)
    print("[INFO] continuous outcome:", OUTCOME_CONT)


if __name__ == "__main__":
    main()
